#!/usr/bin/env python3
"""
Excel File Analyzer for Liasse Fiscale
Analyzes Excel structure to help create React components
"""

import sys
import json
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Alignment
import pandas as pd

def analyze_excel_file(file_path):
    """
    Analyze Excel file and extract comprehensive structure information
    """
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        analysis = {
            "file_info": {
                "filename": Path(file_path).name,
                "total_sheets": len(workbook.sheetnames)
            },
            "sheets": []
        }
        
        # Analyze each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_analysis = analyze_sheet(sheet, sheet_name)
            analysis["sheets"].append(sheet_analysis)
        
        return analysis
        
    except Exception as e:
        return {"error": f"Failed to analyze Excel file: {str(e)}"}

def analyze_sheet(sheet, sheet_name):
    """
    Analyze individual sheet structure
    """
    sheet_info = {
        "name": sheet_name,
        "dimensions": {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "max_column_letter": get_column_letter(sheet.max_column)
        },
        "tables": [],
        "merged_cells": [],
        "styling": {
            "color_patterns": [],
            "font_patterns": [],
            "border_patterns": []
        },
        "formulas": [],
        "headers": [],
        "structure_analysis": {}
    }
    
    # Get merged cells
    for merged_range in sheet.merged_cells.ranges:
        sheet_info["merged_cells"].append(str(merged_range))
    
    # Analyze cell content and styling
    cells_data = []
    color_map = {}
    font_map = {}
    
    for row in range(1, min(sheet.max_row + 1, 200)):  # Limit to first 200 rows for performance
        for col in range(1, min(sheet.max_column + 1, 50)):  # Limit to first 50 columns
            cell = sheet.cell(row=row, column=col)
            
            if cell.value is not None:
                cell_info = {
                    "address": cell.coordinate,
                    "row": row,
                    "column": col,
                    "value": str(cell.value) if cell.value is not None else "",
                    "data_type": str(type(cell.value).__name__),
                    "has_formula": bool(cell.value and str(cell.value).startswith('=')),
                    "styling": {}
                }
                
                # Extract styling information
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    fill_color = cell.fill.start_color.rgb
                    cell_info["styling"]["fill_color"] = fill_color
                    if fill_color not in color_map:
                        color_map[fill_color] = []
                    color_map[fill_color].append(cell.coordinate)
                
                if cell.font:
                    font_info = {
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "size": cell.font.size,
                        "name": cell.font.name
                    }
                    if cell.font.color and cell.font.color.rgb:
                        font_info["color"] = cell.font.color.rgb
                    
                    cell_info["styling"]["font"] = font_info
                    
                    font_key = f"{font_info.get('bold', False)}_{font_info.get('size', 11)}_{font_info.get('name', 'Calibri')}"
                    if font_key not in font_map:
                        font_map[font_key] = []
                    font_map[font_key].append(cell.coordinate)
                
                if cell.border and (cell.border.left.style or cell.border.right.style or 
                                  cell.border.top.style or cell.border.bottom.style):
                    cell_info["styling"]["has_border"] = True
                
                if cell.alignment:
                    alignment_info = {
                        "horizontal": cell.alignment.horizontal,
                        "vertical": cell.alignment.vertical,
                        "wrap_text": cell.alignment.wrap_text
                    }
                    cell_info["styling"]["alignment"] = alignment_info
                
                # Check if it's a formula
                if cell.value and str(cell.value).startswith('='):
                    sheet_info["formulas"].append({
                        "address": cell.coordinate,
                        "formula": str(cell.value)
                    })
                
                cells_data.append(cell_info)
    
    # Store color and font patterns
    sheet_info["styling"]["color_patterns"] = [
        {"color": color, "cells": cells} for color, cells in color_map.items()
    ]
    sheet_info["styling"]["font_patterns"] = [
        {"pattern": pattern, "cells": cells} for pattern, cells in font_map.items()
    ]
    
    # Identify potential tables and headers
    sheet_info["tables"] = identify_tables(cells_data)
    sheet_info["headers"] = identify_headers(cells_data)
    sheet_info["structure_analysis"] = analyze_structure_patterns(cells_data, sheet)
    
    return sheet_info

def identify_tables(cells_data):
    """
    Identify table structures in the sheet
    """
    tables = []
    
    # Group cells by row to identify table-like structures
    rows_data = {}
    for cell in cells_data:
        row = cell["row"]
        if row not in rows_data:
            rows_data[row] = []
        rows_data[row].append(cell)
    
    # Look for consecutive rows with similar column patterns
    consecutive_rows = []
    prev_cols = set()
    
    for row_num in sorted(rows_data.keys()):
        current_cols = set(cell["column"] for cell in rows_data[row_num])
        
        if len(current_cols) >= 3:  # At least 3 columns to be considered a table
            if prev_cols and len(current_cols.intersection(prev_cols)) >= 2:
                # Similar column structure, likely part of same table
                if not consecutive_rows:
                    consecutive_rows = [row_num - 1, row_num]
                else:
                    consecutive_rows.append(row_num)
            else:
                # Process previous table if exists
                if len(consecutive_rows) >= 3:
                    tables.append({
                        "start_row": consecutive_rows[0],
                        "end_row": consecutive_rows[-1],
                        "columns": sorted(prev_cols),
                        "estimated_header_row": consecutive_rows[0]
                    })
                consecutive_rows = [row_num]
            
            prev_cols = current_cols
    
    # Don't forget the last table
    if len(consecutive_rows) >= 3:
        tables.append({
            "start_row": consecutive_rows[0],
            "end_row": consecutive_rows[-1],
            "columns": sorted(prev_cols),
            "estimated_header_row": consecutive_rows[0]
        })
    
    return tables

def identify_headers(cells_data):
    """
    Identify header rows and cells based on styling
    """
    headers = []
    
    for cell in cells_data:
        # Check for header characteristics
        is_header = False
        reasons = []
        
        styling = cell.get("styling", {})
        font = styling.get("font", {})
        
        if font.get("bold"):
            is_header = True
            reasons.append("bold_font")
        
        if styling.get("fill_color") and styling["fill_color"] != "FFFFFF":
            is_header = True
            reasons.append("background_color")
        
        if styling.get("has_border"):
            reasons.append("has_border")
        
        # Check if value looks like a header (short, descriptive text)
        value = str(cell["value"]).strip()
        if len(value) < 50 and not value.replace(".", "").replace(",", "").replace(" ", "").isdigit():
            if any(keyword in value.lower() for keyword in 
                   ["total", "titre", "designation", "montant", "exercice", "compte", "libelle"]):
                is_header = True
                reasons.append("header_keywords")
        
        if is_header:
            headers.append({
                "address": cell["address"],
                "row": cell["row"],
                "column": cell["column"],
                "value": cell["value"],
                "reasons": reasons
            })
    
    return headers

def analyze_structure_patterns(cells_data, sheet):
    """
    Analyze structural patterns in the sheet
    """
    patterns = {
        "title_cells": [],
        "calculation_cells": [],
        "input_cells": [],
        "total_rows": [],
        "section_dividers": []
    }
    
    for cell in cells_data:
        value = str(cell["value"]).strip()
        styling = cell.get("styling", {})
        
        # Title cells (usually bold, larger font, or center-aligned)
        if styling.get("font", {}).get("bold") and len(value) > 0:
            if cell["row"] <= 10:  # Usually titles are in first 10 rows
                patterns["title_cells"].append(cell)
        
        # Calculation cells (contain formulas)
        if cell.get("has_formula"):
            patterns["calculation_cells"].append(cell)
        
        # Total rows (contain "total" keyword)
        if "total" in value.lower():
            patterns["total_rows"].append(cell)
        
        # Section dividers (empty rows or rows with only one cell)
        if value == "" and styling.get("fill_color"):
            patterns["section_dividers"].append(cell)
    
    return patterns

def main():
    if len(sys.argv) != 2:
        print("Usage: python excel_analyzer.py <excel_file_path>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not Path(file_path).exists():
        print(f"Error: File '{file_path}' does not exist")
        sys.exit(1)
    
    print("Analyzing Excel file...")
    analysis = analyze_excel_file(file_path)
    
    if "error" in analysis:
        print(f"Error: {analysis['error']}")
        sys.exit(1)
    
    # Pretty print the analysis
    print("\n" + "="*80)
    print(f"EXCEL FILE ANALYSIS: {analysis['file_info']['filename']}")
    print("="*80)
    
    print(f"\nTotal Sheets: {analysis['file_info']['total_sheets']}")
    
    for sheet in analysis["sheets"]:
        print(f"\n{'='*60}")
        print(f"SHEET: {sheet['name']}")
        print(f"{'='*60}")
        
        dims = sheet['dimensions']
        print(f"Dimensions: {dims['max_row']} rows × {dims['max_column']} columns ({dims['max_column_letter']})")
        
        if sheet['merged_cells']:
            print(f"\nMerged Cells ({len(sheet['merged_cells'])}): {', '.join(sheet['merged_cells'][:10])}")
            if len(sheet['merged_cells']) > 10:
                print(f"... and {len(sheet['merged_cells']) - 10} more")
        
        if sheet['headers']:
            print(f"\nIdentified Headers ({len(sheet['headers'])}):")
            for header in sheet['headers'][:10]:
                print(f"  - {header['address']}: {header['value']} ({', '.join(header['reasons'])})")
            if len(sheet['headers']) > 10:
                print(f"  ... and {len(sheet['headers']) - 10} more")
        
        if sheet['tables']:
            print(f"\nIdentified Tables ({len(sheet['tables'])}):")
            for i, table in enumerate(sheet['tables'], 1):
                print(f"  Table {i}: Rows {table['start_row']}-{table['end_row']}, Columns: {table['columns']}")
        
        if sheet['formulas']:
            print(f"\nFormulas ({len(sheet['formulas'])}):")
            for formula in sheet['formulas'][:5]:
                print(f"  - {formula['address']}: {formula['formula']}")
            if len(sheet['formulas']) > 5:
                print(f"  ... and {len(sheet['formulas']) - 5} more")
        
        color_patterns = sheet['styling']['color_patterns']
        if color_patterns:
            print(f"\nColor Patterns ({len(color_patterns)}):")
            for pattern in color_patterns[:5]:
                print(f"  - Color {pattern['color']}: {len(pattern['cells'])} cells")
            if len(color_patterns) > 5:
                print(f"  ... and {len(color_patterns) - 5} more colors")
        
        structure = sheet['structure_analysis']
        if structure['title_cells']:
            print(f"\nTitle Cells: {len(structure['title_cells'])}")
        if structure['calculation_cells']:
            print(f"Calculation Cells: {len(structure['calculation_cells'])}")
        if structure['total_rows']:
            print(f"Total Rows: {len(structure['total_rows'])}")
    
    # Save detailed analysis to JSON file
    output_file = Path(file_path).stem + "_analysis.json"
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(analysis, f, indent=2, ensure_ascii=False)
        print(f"\n\nDetailed analysis saved to: {output_file}")
    except Exception as e:
        print(f"\nWarning: Could not save analysis to file: {e}")

if __name__ == "__main__":
    main()